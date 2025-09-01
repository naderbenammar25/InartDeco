import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_excel_template():
    """
    Génère un fichier Excel template avec les colonnes correspondant aux modèles Django
    """
    
    # Définir les colonnes pour chaque table selon vos modèles Django
    
    # Table Catégories
    categories_columns = [
        'nom',
        'description', 
        'parent_id',  # ID de la catégorie parent (optionnel)
        'image',
        'active',
        'meta_title',
        'meta_description'
    ]
    
    # Table Marques
    marques_columns = [
        'nom',
        'description',
        'logo',
        'site_web',
        'active'
    ]
    
    # Table Fournisseurs
    fournisseurs_columns = [
        'nom',
        'description',
        'contact_nom',
        'contact_email',
        'contact_telephone',
        'adresse',
        'ville',
        'code_postal',
        'pays',
        'site_web',
        'active'
    ]
    
    # Table Produits (colonnes principales)
    produits_columns = [
        'nom',
        'description',
        'prix',
        'prix_promo',
        'stock',
        'seuil_stock',
        'categorie_nom',  # Nom de la catégorie (sera mappé avec l'ID)
        'marque_nom',     # Nom de la marque (sera mappé avec l'ID)
        'fournisseur_nom', # Nom du fournisseur (sera mappé avec l'ID)
        'reference',
        'sku',
        'code_barre',
        'poids',
        'dimensions',
        'couleur',
        'materiau',
        'etat',
        'image_principale',
        'active',
        'featured',
        'nouveau',
        'meta_title',
        'meta_description'
    ]
    
    # Table Images supplémentaires
    images_columns = [
        'produit_reference',  # Référence du produit
        'image_url',
        'alt_text',
        'ordre'
    ]
    
    # Créer le workbook Excel
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    wb.remove(wb.active)
    
    # Styles pour l'en-tête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def create_sheet_with_data(sheet_name, columns, example_data=None):
        """Créer une feuille avec les colonnes et optionnellement des données d'exemple"""
        ws = wb.create_sheet(title=sheet_name)
        
        # Ajouter les en-têtes
        for col, column_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=column_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Ajouter des données d'exemple si fournies
        if example_data:
            for row_idx, row_data in enumerate(example_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return ws
    
    # Données d'exemple pour Catégories
    categories_examples = [
        ["Cuisine", "Équipements et ustensiles de cuisine", "", "", "True", "Cuisine - InArtDeco", "Découvrez notre gamme cuisine"],
        ["Électroménager", "Appareils électriques pour la cuisine", "1", "", "True", "Électroménager", "Robots, mixeurs, etc."],
        ["Salon", "Mobilier et décoration salon", "", "", "True", "Salon - InArtDeco", "Canapés, tables, déco salon"],
        ["Éclairage", "Luminaires et éclairage", "", "", "True", "Éclairage", "Suspensions, lampes, etc."]
    ]
    
    # Données d'exemple pour Marques
    marques_examples = [
        ["KitchenAid", "Électroménager haut de gamme", "", "https://www.kitchenaid.com", "True"],
        ["Ikea", "Mobilier et décoration design", "", "https://www.ikea.com", "True"],
        ["Le Creuset", "Batterie de cuisine premium", "", "https://www.lecreuset.com", "True"],
        ["Maisons du Monde", "Décoration et mobilier", "", "https://www.maisonsdumonde.com", "True"]
    ]
    
    # Données d'exemple pour Fournisseurs
    fournisseurs_examples = [
        ["Fournisseur Cuisine Pro", "Spécialiste électroménager", "Ahmed Ben Ali", "ahmed@cuisinepro.tn", "+216 71 123 456", "Rue de la République", "Tunis", "1000", "Tunisie", "www.cuisinepro.tn", "True"],
        ["Mobilier Design", "Import mobilier européen", "Fatma Gharbi", "contact@mobilierdesign.tn", "+216 70 987 654", "Avenue Habib Bourguiba", "Sousse", "4000", "Tunisie", "", "True"]
    ]
    
    # Données d'exemple pour Produits
    produits_examples = [
        ["Robot pâtissier KitchenAid Artisan", "Robot pâtissier professionnel avec bol inox 4,8L", "1450.00", "1299.00", "15", "3", "Électroménager", "KitchenAid", "Fournisseur Cuisine Pro", "KA-ART-001", "KITCHENAID-5KSM175PS", "", "11.5", "37 x 24 x 36 cm", "Rouge Empire", "Métal", "neuf", "", "True", "True", "True", "Robot KitchenAid Artisan", "Robot pâtissier professionnel"],
        ["Canapé 3 places Björk", "Canapé scandinave tissu gris", "1899.00", "1599.00", "5", "1", "Salon", "Maisons du Monde", "Mobilier Design", "MDM-CAN-001", "BJORK-3P-GREY", "", "45.0", "190 x 85 x 78 cm", "Gris chiné", "Tissu/Bois", "neuf", "", "True", "True", "False", "Canapé Scandinave Björk", "Canapé 3 places design"],
        ["Plat gratin Pyrex 35cm", "Plat à gratin verre borosilicate", "79.90", "", "25", "5", "Cuisine", "Pyrex", "Fournisseur Cuisine Pro", "PY-GRAT-001", "PYREX-GRAT35", "", "1.8", "35 x 23 x 6 cm", "Transparent", "Verre", "neuf", "", "True", "False", "False", "Plat Gratin Pyrex", "Plat gratin verre résistant"]
    ]
    
    # Données d'exemple pour Images
    images_examples = [
        ["KA-ART-001", "https://example.com/kitchenaid1.jpg", "Robot KitchenAid vue face", "1"],
        ["KA-ART-001", "https://example.com/kitchenaid2.jpg", "Robot KitchenAid vue profil", "2"],
        ["MDM-CAN-001", "https://example.com/canape1.jpg", "Canapé Björk vue face", "1"]
    ]
    
    # Créer les feuilles
    create_sheet_with_data("1-Catégories", categories_columns, categories_examples)
    create_sheet_with_data("2-Marques", marques_columns, marques_examples)
    create_sheet_with_data("3-Fournisseurs", fournisseurs_columns, fournisseurs_examples)
    create_sheet_with_data("4-Produits", produits_columns, produits_examples)
    create_sheet_with_data("5-Images", images_columns, images_examples)
    
    # Créer une feuille d'instructions
    instructions_ws = wb.create_sheet("0-Instructions", 0)
    instructions = [
        ["INSTRUCTIONS POUR REMPLIR LE FICHIER EXCEL"],
        [""],
        ["1. ORDRE DE REMPLISSAGE (IMPORTANT):"],
        ["   - Commencez par la feuille '1-Catégories'"],
        ["   - Puis '2-Marques'"],
        ["   - Ensuite '3-Fournisseurs'"],
        ["   - Puis '4-Produits'"],
        ["   - Enfin '5-Images' (optionnel)"],
        [""],
        ["2. CONSEILS POUR CHAQUE FEUILLE:"],
        [""],
        ["CATÉGORIES:"],
        ["- nom: Nom unique de la catégorie"],
        ["- parent_id: Laissez vide pour catégorie principale, sinon mettez l'ID de la catégorie parent"],
        ["- active: True ou False"],
        [""],
        ["MARQUES:"],
        ["- nom: Nom unique de la marque"],
        ["- site_web: URL complète (optionnel)"],
        [""],
        ["FOURNISSEURS:"],
        ["- Remplissez au minimum: nom, contact_nom, contact_email"],
        [""],
        ["PRODUITS:"],
        ["- categorie_nom: Doit correspondre exactement au nom d'une catégorie créée"],
        ["- marque_nom: Doit correspondre exactement au nom d'une marque créée"],
        ["- fournisseur_nom: Doit correspondre exactement au nom d'un fournisseur créé"],
        ["- reference: Référence unique obligatoire"],
        ["- prix: Format décimal (ex: 1450.00)"],
        ["- etat: neuf, occasion, ou reconditionne"],
        ["- featured/nouveau/active: True ou False"],
        [""],
        ["IMAGES:"],
        ["- produit_reference: Doit correspondre à une référence de produit"],
        ["- ordre: Numéro d'ordre d'affichage (1, 2, 3...)"],
        [""],
        ["3. FORMATS IMPORTANTS:"],
        ["- Prix: Format décimal avec point (1450.00)"],
        ["- Booléens: True ou False (respecter la casse)"],
        ["- Dates: YYYY-MM-DD"],
        ["- URLs: Commencer par http:// ou https://"],
        [""],
        ["4. CHAMPS OBLIGATOIRES:"],
        ["- Produits: nom, prix, stock, categorie_nom, reference"],
        ["- Catégories: nom"],
        ["- Marques: nom"],
        [""],
        ["5. APRÈS REMPLISSAGE:"],
        ["- Sauvegardez le fichier Excel"],
        ["- Envoyez-le au développeur pour import en base de données"],
    ]
    
    for row_idx, instruction in enumerate(instructions, 1):
        cell = instructions_ws.cell(row=row_idx, column=1, value=instruction[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=14)
        elif instruction[0].endswith(":") and not instruction[0].startswith("   "):
            cell.font = Font(bold=True, color="366092")
    
    instructions_ws.column_dimensions['A'].width = 80
    
    # Sauvegarder le fichier
    filename = "template_produits_inartdeco.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)
    
    return filepath

if __name__ == "__main__":
    try:
        filepath = create_excel_template()
        print(f"✅ Fichier Excel template créé avec succès!")
        print(f"📁 Emplacement: {filepath}")
        print(f"📋 Le fichier contient:")
        print(f"   - Feuille Instructions")
        print(f"   - Feuille Catégories (avec exemples)")
        print(f"   - Feuille Marques (avec exemples)")
        print(f"   - Feuille Fournisseurs (avec exemples)")
        print(f"   - Feuille Produits (avec exemples)")
        print(f"   - Feuille Images (avec exemples)")
        print(f"\n💡 Votre client peut maintenant:")
        print(f"   1. Supprimer les exemples")
        print(f"   2. Remplir avec ses vrais produits")
        print(f"   3. Vous renvoyer le fichier pour import")
        
    except Exception as e:
        print(f"❌ Erreur lors de la création du fichier: {str(e)}")